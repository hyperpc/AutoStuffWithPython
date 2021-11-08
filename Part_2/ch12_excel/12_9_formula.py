import openpyxl, sys, os

currentPath = os.path.dirname(sys.argv[0])
excelPath = os.path.join(currentPath, "writeFormula.xlsx")

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(1,2).value = 82
sheet.cell(2,2).value = 11
sheet.cell(3,2).value = 85
sheet.cell(4,2).value = 18
sheet.cell(5,2).value = 57
sheet.cell(6,2).value = 51
sheet.cell(7,2).value = 38
sheet.cell(8,2).value = 42
sheet.cell(9,1).value = 'TOTAL: '
sheet.cell(9,2).value = '=SUM(B1:B8)'
wb.save(excelPath)

wbFormulas = openpyxl.load_workbook(excelPath)
activeSheet_WF = wbFormulas.active
print('sheet[\'B9\'].value(default): ', activeSheet_WF.cell(9,2).value)
# here is a bug: need save mannually firstly, then load wb with data_only=True
wbDataOnly = openpyxl.load_workbook(excelPath, data_only=True)
activeSheet_WD = wbDataOnly.active
print('sheet[\'B9\'].value(data_only): ', activeSheet_WD.cell(9,2).value)
