import openpyxl
import sys, os
from openpyxl.styles import Font

currentPath = os.path.dirname(sys.argv[0])
excelPath = os.path.join(currentPath, "styles.xlsx")

wb = openpyxl.Workbook()
sheet = wb['Sheet']
font1 = Font(name='Times New Roman', bold=True)
sheet.cell(1, 1).font = font1
sheet.cell(1, 1).value = 'Bold Times New Roman'
font2 = Font(size=24, italic=True)
sheet.cell(3, 2).font = font2
sheet.cell(3, 2).value = '24 pt Italic'
wb.save(excelPath)