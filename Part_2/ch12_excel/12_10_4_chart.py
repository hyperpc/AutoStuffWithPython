import sys, os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from copy import deepcopy
from openpyxl import load_workbook

def drawBarCharts():
    currentPath = os.path.dirname(sys.argv[0])
    excelPath = os.path.join(currentPath, "barChart.xlsx")
    excelPath_Copy = os.path.join(currentPath, "barChart_copy.xlsx")
    
    '''
    wb = Workbook()
    activeSheet = wb.active
    for i in range(1, 11): # create some data in column 'A'
        activeSheet.cell(i, 1).value = i
    refObj = Reference(activeSheet, 1, 1, 10, 1)
    seriesObj = Series(refObj, title = 'First series')
    chartObj = BarChart()
    chartObj.append(seriesObj)
    chartObj.drawing.top = 50 # set the position
    chartObj.drawing.left = 100 # set the position
    chartObj.drawing.width = 300 # set the size
    chartObj.drawing.height = 200 # set the size
    activeSheet.add_chart(chartObj)
    wb.save(excelPath)
    '''
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    rows = [
        ('Number', 'Batch 1', 'Batch 2')
        ,(2, 10, 30)
        ,(3, 40, 60)
        ,(4, 50, 70)
        ,(5, 20, 10)
        ,(6, 10, 40)
        ,(7, 50, 30)
    ]
    for row in rows:
        ws.append(row)

    chartObj = BarChart()
    chartObj.style = 10
    chartObj.type = 'col'
    chartObj.title = 'Bar Chart'
    chartObj.y_axis.title = 'Test number'
    chartObj.x_axis.title = 'Sample length(mm)'

    data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chartObj.add_data(data, titles_from_data=True)
    chartObj.set_categories(cats)
    chartObj.shape = 4
    ws.add_chart(chartObj, 'A10')

    chartObj2 = deepcopy(chartObj)
    chartObj2.style = 11
    chartObj2.type = 'bar'
    chartObj2.title = 'Horizontal Bar Chart'
    ws.add_chart(chartObj2, 'J10')

    chartObj3 = deepcopy(chartObj)
    chartObj3.style = 12
    chartObj3.type = 'col'
    chartObj3.grouping = 'stacked'
    chartObj3.overlap = 100
    chartObj3.title = 'Stacked Chart'
    ws.add_chart(chartObj3, 'A27')

    chartObj4 = deepcopy(chartObj)
    chartObj4.style = 13
    chartObj4.type = 'bar'
    chartObj4.grouping = 'percentStacked'
    chartObj4.overlap = 100
    chartObj4.title = 'Percent Stacked Chart'
    ws.add_chart(chartObj4, 'J27')

    wb.save(excelPath)

    wb = load_workbook(excelPath)
    wb.save(excelPath_Copy)

def main():
    drawBarCharts()

if __name__ == '__main__':
    main()
