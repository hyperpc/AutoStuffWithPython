import openpyxl, sys, os

currentPath = os.path.dirname(sys.argv[0])
excelPath = os.path.join(currentPath, "dimension.xlsx")

wb = openpyxl.Workbook()
activeSheet = wb.active
activeSheet.cell(1, 1).value = 'Tall row'
activeSheet.cell(2, 2).value = 'Wide column'
activeSheet.row_dimensions[1].height = 70
activeSheet.column_dimensions['B'].width = 20
wb.save(excelPath)