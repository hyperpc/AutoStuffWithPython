import sys, os
from openpyxl.cell import cell
from openpyxl.styles.colors import Color
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NumberFormatDescriptor

def reverseRowCol():    
    currentPath = os.path.dirname(sys.argv[0])
    excelPath_org = os.path.join(currentPath, 'example.xlsx')
    excelPath_new = os.path.join(currentPath, 'example_reversed.xlsx')

    wb_org = load_workbook(excelPath_org)
    activeSheet_org = wb_org.active
    maxRow = activeSheet_org.max_row
    maxCol = activeSheet_org.max_column
    sheetData = []
    for i in range(1, maxRow + 1):
        rowData = []
        for j in range(1, maxCol + 1):
            rowData.append(activeSheet_org.cell(i, j).value)
        sheetData.append(rowData)
    
    wb_new = Workbook()
    activeSheet_new = wb_new.active
    for i in range(1, maxCol + 1):
        for j in range(1, maxRow + 1):
            activeSheet_new.cell(i,j).value = sheetData[j-1][i-1]
    wb_new.save(excelPath_new)
    print('Done!')

def reverseRowColStyled():
    currentPath = os.path.dirname(sys.argv[0])
    excelPath_org = os.path.join(currentPath, 'example.xlsx')
    excelPath_new = os.path.join(currentPath, 'example_reversed_styled.xlsx')

    wb_new = Workbook()
    activeSheet_new = wb_new.active

    wb_org = load_workbook(excelPath_org)
    activeSheet_org = wb_org.active
    maxRow = activeSheet_org.max_row
    maxCol = activeSheet_org.max_column
    maxRow_height = 15
    maxCol_width = 30
    for i in range(1, maxRow + 1):
        if activeSheet_org.row_dimensions[i].height != None and activeSheet_org.row_dimensions[i].height > maxRow_height:
            maxRow_height = activeSheet_org.row_dimensions[i].height
        else:
            maxRow_height = 15
        for j in range(1, maxCol + 1):
            col_letter = cell.get_column_letter(j)
            if activeSheet_org.column_dimensions[col_letter].width != None and activeSheet_org.column_dimensions[col_letter].width > maxCol_width:
                maxCol_width = activeSheet_org.column_dimensions[col_letter].width
            else:
                maxCol_width = 30

            cell_org = activeSheet_org.cell(i, j)
            cell_new = activeSheet_new.cell(j, i)
            cell_new.value = cell_org.value
            
            cell_new.data_type = cell_org.data_type
            fill = PatternFill(
                fill_type=cell_org.fill.fill_type
                ,fgColor=cell_org.fill.fgColor.rgb
                ,bgColor=cell_org.fill.bgColor.rgb
            )
            cell_new.fill = fill
            #if cell_org.has_style:
            font_color = '00000000'
            if isinstance(cell_org.font.color.rgb, str) and isinstance(cell_org.font.color.index, str):
                font_color = cell_org.font.color.rgb
            font = Font(
                name=cell_org.font.name
                ,size=cell_org.font.sz
                ,bold=cell_org.font.b
                ,italic=cell_org.font.italic
                ,vertAlign=cell_org.font.vertAlign
                ,underline=cell_org.font.underline
                ,strike=cell_org.font.strike
                ,color=font_color
            )
            cell_new.font = font
            border = Border(
                left=Side(border_style=cell_org.border.left.border_style, color=cell_org.border.left.color)
                ,right=Side(border_style=cell_org.border.right.border_style, color=cell_org.border.right.color)
                ,top=Side(border_style=cell_org.border.top.border_style, color=cell_org.border.top.color)
                ,bottom=Side(border_style=cell_org.border.bottom.border_style, color=cell_org.border.bottom.color)
                ,diagonal=Side(border_style=cell_org.border.diagonal.border_style, color=cell_org.border.diagonal.color)
                ,diagonal_direction=cell_org.border.diagonal_direction
                ,outline=cell_org.border.outline
            )
            cell_new.border = border
            protection = Protection(
                locked=cell_org.protection.locked
                ,hidden=cell_org.protection.hidden
            )
            cell_new.protection = protection
            alignment = Alignment(
                horizontal=cell_org.alignment.horizontal
                ,vertical=cell_org.alignment.vertical
                ,text_rotation=cell_org.alignment.text_rotation
                ,wrap_text=cell_org.alignment.wrap_text
                ,shrink_to_fit=cell_org.alignment.shrink_to_fit
                ,indent=cell_org.alignment.indent
            )
            cell_new.alignment = alignment
            cell_new.number_format = cell_org.number_format
            if cell_org.hyperlink:
                cell_new._hyperlink = cell_org.hyperlink
            if cell_org.comment:
                cell_new.comment = cell_org.comment

    for i in range(1, maxCol + 1):
        activeSheet_new.row_dimensions[i].hieght = maxRow_height
        for j in range(1, maxRow + 1):
            col_letter = cell.get_column_letter(j)
            activeSheet_new.column_dimensions[col_letter].width = maxCol_width
    wb_new.save(excelPath_new)
    print('Done!')
    
def main():
    #reverseRowCol()
    reverseRowColStyled()

if __name__ == '__main__':
    main()