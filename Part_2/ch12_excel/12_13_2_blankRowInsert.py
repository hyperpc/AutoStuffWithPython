#! python3
import sys, os
from openpyxl.cell.cell import Cell
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# won't lost style and data format
def insertBlankRow_v1():
    n = 3
    m = 2
    filename = 'myProduce_v1.xlsx'

    try:
        n = int(sys.argv[1])
    except Exception as ex:
        print('Exception: ', ex)
        n = 3
    
    try:
        m = int(sys.argv[2])
    except Exception as ex:
        print('Exception: ', ex)
        m = 2
    
    filename = sys.argv[3]
    if filename == None:
        filename = 'myProduce_v1.xlsx'
    elif filename == '':
        filename = 'myProduce_v1.xlsx'
    elif filename == 'myProduce.xlsx':
        filename = 'myProduce_v1.xlsx'
    
    currentPath = os.path.dirname(sys.argv[0])
    excelPath_org = os.path.join(currentPath, 'example.xlsx')
    excelPath_new = os.path.join(currentPath, filename)

    wb_org = load_workbook(excelPath_org)
    ws_org = wb_org.active
    ws_org.insert_rows(n, m)
    wb_org.save(excelPath_new)

# this method will lost the style and data format
def insertBlankRow_v2():
    n = 3
    m = 2
    filename = 'myProduce_v2.xlsx'

    try:
        n = int(sys.argv[1])
    except Exception as ex:
        print('Exception: ', ex)
        n = 3
    
    try:
        m = int(sys.argv[2])
    except Exception as ex:
        print('Exception: ', ex)
        m = 2
    
    filename = sys.argv[3]
    if filename == None:
        filename = 'myProduce_v2.xlsx'
    elif filename == '':
        filename = 'myProduce_v2.xlsx'
    elif filename == 'myProduce.xlsx':
        filename = 'myProduce_v2.xlsx'
    
    currentPath = os.path.dirname(sys.argv[0])
    excelPath_org = os.path.join(currentPath, 'example.xlsx')
    excelPath_new = os.path.join(currentPath, filename)

    wb_org = load_workbook(excelPath_org, data_only=True)
    ws_org = wb_org.active
    wb_new = Workbook()
    ws_new = wb_new.active

    # range counter rows/columns of org file
    '''
    row_counter = 0
    for row in ws_org.iter_rows():
        col_counter = 0
        for col in ws_org.iter_cols():
            col_counter += 1
        row_counter += 1
    '''
    row_org_counter = ws_org.max_row
    col_org_counter = ws_org.max_column

    # copy logic
    rowList = [] # list of lists of rows
    for i in range(1, row_org_counter + m + 1):
        rowSelected = [] # list of row
        for j in range(1, col_org_counter + 1):
            rowSelected.append(ws_org.cell(i, j).value) # cell values are loaded into list of rows
        rowList.append(rowSelected)
    #print(rowList)

    # range counter rows/columns of new file
    row_new_counter = row_org_counter + m
    col_new_counter = col_org_counter
    # paste logic
    counterRow = 0
    for i in range(1, row_new_counter + 1):
        counterCol = 0
        for j in range(1, col_new_counter + 1):
            if i < n:
                ws_new.cell(i,j).value = rowList[counterRow][counterCol]
            elif i >= n + m:
                ws_new.cell(i,j).value = rowList[counterRow][counterCol]
            else:
                ws_new.cell(i,j).value = ''
            counterCol += 1
        if i < n:
            counterRow += 1
        elif i >= n + m:
            counterRow += 1
    wb_new.save(excelPath_new)

def main():
    insertBlankRow_v1()
    insertBlankRow_v2()

if __name__ == '__main__':
    main()