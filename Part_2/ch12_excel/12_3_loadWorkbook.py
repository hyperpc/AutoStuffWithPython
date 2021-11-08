import sys, os
from openpyxl import load_workbook
from openpyxl import Workbook as Wb
from openpyxl.cell.cell import Cell
from openpyxl.utils import cell

currentPath = os.path.dirname(sys.argv[0])
excelPath = os.path.join(currentPath, "example.xlsx")
print('excelPath = ',excelPath)
print('')
print('### 12.3.1 load workbook #####')
wb = load_workbook(excelPath)
print('type(wb) = ', type(wb))

print('')
print('### 12.3.2 load sheet #####')
#print('wb.get_sheet_names() = ', wb.get_sheet_names())
print('wb.sheetnames = ', wb.sheetnames)
sheet3 = wb['Sheet3']
print('sheet3 = ', sheet3)
print('type(sheet3) = ', type(sheet3))
print('sheet3.title = ', sheet3.title)
## not supported function get_sheet_by_name
#anotherSheet = wb.get_active_sheet()
activeSheet = wb.active
print('activeSheet = ', activeSheet)
print('type(activeSheet) = ', type(activeSheet))
print('')
print('##### 12.3.3 load cells #####')
sheet1 = wb['Sheet1']
cell_A1_pre = sheet1['A1']
print('cell_A1_pre = sheet1[\'A1\']')
print('cell_A1_pre = ', cell_A1_pre)
print('cell_A1_pre.value = ', cell_A1_pre.value)
cell_A1_new = sheet1.cell(row=1, column=1)
print('cell_A1_new = sheet1.cell(row=1, column=1)')
print('cell_A1_new = ', cell_A1_new)
print('cell_A1_new.value = ', cell_A1_new.value)

cell_B1_pre = sheet1['B1']
print('cell_B1_pre = sheet1[\'B1\']')
print('cell_B1_pre = ', cell_B1_pre)
print('cell_B1_pre.value =', cell_B1_pre.value)
cell_B1_new = sheet1.cell(row=1, column=2)
print('cell_B1_new = sheet1.cell(row=1, column=2)')
print('cell_B1_new = ', cell_B1_new)
print('cell_B1_new.value = ', cell_B1_new.value)
print('Row ', str(cell_B1_new.row), ', Column ', str(cell_B1_new.column), ' is ', cell_B1_new.value)
print('Cell ', cell_B1_new.coordinate, ' is ', cell_B1_new.value)
cell_C1_new = sheet1.cell(row=1, column=3)
print('cell_C1_new = sheet1.cell(row=1, column=3)')
print('cell_C1_new.value = ', cell_C1_new.value)

for i_row in range(1, 8, 2):
    print(i_row, ' ', sheet1.cell(1, 2).value)

print('sheet1.min_row = ', sheet1.min_row)
print('sheet1.min_column = ', sheet1.min_column)
print('sheet1.max_row = ', sheet1.max_row)
print('sheet1.max_column = ', sheet1.max_column)

print('')
print('##### 12.3.4 switch column index and column name #####')
myCell = sheet1.cell(row=1, column=1)
print('myCell.col_idx = ', myCell.col_idx)
print('myCell.column_letter = ', myCell.column_letter)
myCell = sheet1.cell(row=1, column=2)
print('myCell.col_idx = ', myCell.col_idx)
print('myCell.column_letter = ', myCell.column_letter)
myCell = sheet1.cell(row=1, column=27)
print('myCell.col_idx = ', myCell.col_idx)
print('myCell.column_letter = ', myCell.column_letter)
myCell = sheet1.cell(row=1, column=900)
print('myCell.col_idx = ', myCell.col_idx)
print('myCell.column_letter = ', myCell.column_letter)

print('')
print('##### 12.3.5 read row and column from sheet #####')
cellRange = sheet1['A1':'C3']
print('tuple(cellRange) = ', tuple(cellRange))
for rowOfCellObjs in cellRange:
    for cellObj in rowOfCellObjs:
        #print('cellObj.coordinate = ', cellObj.coordinate, ', cellObj.value = ', cellObj.value)
        print(cellObj.coordinate, ' ', cellObj.value)
    print('---End of Row---')
activeSheet = wb.active
print('activeSheet.columns = ', activeSheet.columns)
myColumns = activeSheet.iter_cols(2, 2, 1, 7, True)
for cellObj in myColumns:
    print('cellObj = ', cellObj)
    for item in cellObj:
        print(item)
