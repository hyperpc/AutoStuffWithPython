import openpyxl, sys, os

def mergeCells():
    currentPath = os.path.dirname(sys.argv[0])
    excelPath = os.path.join(currentPath, "merged.xlsx")

    wb = openpyxl.Workbook()
    activeSheet = wb.active
    activeSheet.merge_cells('A1:D3')
    activeSheet.cell(1,1).value = 'Twelve cells merged together'
    activeSheet.merge_cells('C5:D5')
    activeSheet.cell(5,3).value = 'Two merged cells'
    wb.save(excelPath)

def unmergeCells():
    currentPath = os.path.dirname(sys.argv[0])
    excelPath = os.path.join(currentPath, "merged.xlsx")

    wb = openpyxl.load_workbook(excelPath)
    activeSheet = wb.active
    activeSheet.unmerge_cells('A1:D3')
    activeSheet.unmerge_cells('C5:D5')
    wb.save(excelPath)
def main():
    #mergeCells()
    unmergeCells()

if __name__ == '__main__':
    main()