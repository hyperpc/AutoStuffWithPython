import openpyxl
import sys, os
from openpyxl.styles import Font

currentPath = os.path.dirname(sys.argv[0])
excelPath = os.path.join(currentPath, "styled.xlsx")

wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True)
sheet.cell(1, 1).font = italic24Font
sheet.cell(1, 1).value = 'Hello world!'
wb.save(excelPath)