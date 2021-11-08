#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet

import openpyxl, os, sys

# load update data
currentPath = os.path.dirname(sys.argv[0])
excelPath = os.path.join(currentPath, "produceSales.xlsx")
excelPath_update = os.path.join(currentPath, "updatedProduceSales.xlsx")
wb = openpyxl.load_workbook(excelPath)
sheet = wb['Sheet']
# the produce types and their updated prices
PRICE_UPDATES = {
    'Garlic':3.07,
    'Celery':1.19,
    'Lemon':1.27
}

# loop through the rows and update the prices
for rowNum in range(2, sheet.max_row + 1): # skip first row - header row
    produceName = sheet.cell(rowNum, 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(rowNum, 2).value = PRICE_UPDATES[produceName]
    
wb.save(excelPath_update)