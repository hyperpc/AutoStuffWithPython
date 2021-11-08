#! python3

from posixpath import basename
import sys, os
from openpyxl.styles import Font
from openpyxl import Workbook

def multipTable():
    myNum = 0
    while True:
        try:
            print('Enter one number to generate multiplication table: ')
            baseNum = input()
            if baseNum == None:
                continue
            if baseNum == '':
                continue
            if int(baseNum) < 1:
                continue
            myNum = int(baseNum)
            break
        except Exception as ex:
            print('Exception: ', ex)

    currentPath = os.path.dirname(sys.argv[0])
    excelPath = os.path.join(currentPath, "multiplicationTable.xlsx")
    
    wb = Workbook()
    #ws = wb.create_sheet()
    ws = wb.active

    boldFont = Font(bold=True)
    # why myNum + 1 + 1? 
    # - 'A1' is ''
    # - range() exclude the sencond param
    for row_idx in range(1, myNum + 1 + 1): # by columns
        for col_idx in range(1, myNum + 1 + 1): # by rows
            if row_idx == 1:
                if col_idx == 1:
                    ws.cell(row_idx, col_idx).value = ''
                else:
                    ws.cell(row_idx, col_idx).font = boldFont
                    ws.cell(row_idx, col_idx).value = col_idx - 1
            else:
                if col_idx == 1:
                    ws.cell(row_idx, col_idx).font = boldFont
                    ws.cell(row_idx, col_idx).value = row_idx - 1
                else:
                    ws.cell(row_idx, col_idx).value = (row_idx - 1) * (col_idx - 1)
    wb.save(excelPath)

def main():
    multipTable()

if __name__ == '__main__':
    main()