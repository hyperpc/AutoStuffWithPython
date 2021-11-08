import openpyxl, pprint, os, sys

currentPath = os.path.dirname(sys.argv[0])
excelPath = os.path.join(currentPath, "censuspopdata.xlsx")
print('Opening workbook...')
wb = openpyxl.load_workbook(excelPath)
sheet = wb['Population by Census Tract']
countyData = {}

# Fill in countyData with each county's population and tracts.
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # each row in the spreadsheet has data for one census tract.
    State = sheet.cell(row, 2).value
    County = sheet.cell(row, 3).value
    Pop = sheet.cell(row, 4).value
    # make sure the key for this state exists
    countyData.setdefault(State, {})
    # make sure the key for this county in this state exists
    countyData[State].setdefault(County, {'tracts':0, 'pop':0})
    # each row represents one census tract, so increment by one
    countyData[State][County]['tracts'] += 1
    # increase the county pop by the pop in this census tract
    countyData[State][County]['pop'] += int(Pop)
# open a new text file and write the contents of cuntyData to it
print('Writing results...')
resultFile = open(os.path.join(currentPath, "census2010.py"), 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')
