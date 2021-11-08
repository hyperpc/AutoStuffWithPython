import sys, os
from openpyxl import load_workbook
    
def txtToExcel():
    currentPath = os.path.dirname(sys.argv[0])
    txtFiles = ['1_new.txt', '2_new.txt', '3_new.txt']
    excelPath = os.path.join(currentPath, 'txtExcel.xlsx')

    wb = load_workbook(excelPath, data_only=True)
    activeSheet = wb.active

    for i in range(1, activeSheet.max_column + 1):
        txtFilePath = os.path.join(currentPath, txtFiles[i-1])
        txtFile = open(txtFilePath, 'w')
        for j in range(1, activeSheet.max_row + 1):
            txtFile.write(activeSheet.cell(j, i).value)
        txtFile.close()
    print('Done!')

def main():
    txtToExcel()

if __name__ == '__main__':
    main()