#! python3
import openpyxl, smtplib, sys, os

# open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook('duesRecords.xlsx')
ws = wb['Sheet1']
lastCol = ws.max_column()
latestMonth = ws.cell(row=1, column=lastCol).value

# check each member's payment status
unpaidMembers = []
for r in range(2, ws.max_row+1):
    payment = ws.cell(r,lastCol).value
    if payment!='paid':
        name=ws.cell(r,1).value
        email=ws.cell(r,2).value
        unpaidMembers[name] = email

# login to email account
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('email@addr.com', sys.argv[1])

# send out reminder emails
for name, email in unpaidMembers.items():
    body = 'Subject: %s dues unpaid. \nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!' % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendemailStatus = smtpObj.sendmail('from_addr', email, body)
    if sendemailStatus!={}:
        print('There was a problem sending email to %s: %s' % (email, sendemailStatus))
smtpObj.quit()
