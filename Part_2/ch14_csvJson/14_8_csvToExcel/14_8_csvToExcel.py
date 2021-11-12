import openpyxl, sys, os, csv

def csvToExcel():
    currentPath = os.path.dirname(sys.argv[0])
    csvPath = os.path.join(currentPath, 'csv')
    excelPath = os.path.join(currentPath, 'excel')
    filenames = os.listdir(csvPath)
    file_cnt = len(filenames)
    names = filenames[0].split('-')
    excel_basename = names[0]
    excel_fullname = names[0] + '.xlsx'
    wb = openpyxl.Workbook()
    for i in range(0, file_cnt):
        current_csv_name = filenames[i]
        if not current_csv_name.endswith('.csv'):
            continue
        print('>>> Processing file(%d): %s' %(i, current_csv_name))
        sheetname = current_csv_name.replace(excel_basename + '-', '').replace('.csv', '')
        ws = wb.create_sheet(sheetname)

        csvFilePath = os.path.join(csvPath, current_csv_name)
        csvObj = open(csvFilePath)
        reader = csv.reader(csvObj)
        row_idx = 0
        for row in reader:
            col_cnt = len(row)
            for col_idx in range(0, col_cnt):
                ws.cell(row_idx+1, col_idx+1).value = row[col_idx]
            row_idx += 1
        csvObj.close()
    wb.save(os.path.join(excelPath, excel_fullname))
    print('>>> Completed to generate excel file: ', os.path.join(excelPath, excel_fullname))

def main():
    csvToExcel()

if __name__ == '__main__':
    main()