import openpyxl, sys, os, csv

def csvToExcel():
    currentPath = os.path.dirname(sys.argv[0])
    csvPath = os.path.join(currentPath, 'csv')
    excelPath = os.path.join(currentPath, 'excel')
    for excelFile in os.listdir(excelPath):
        if not excelFile.endswith('.xlsx'):
            continue
        print('>>> Poccessing file: ', excelFile)
        wb = openpyxl.load_workbook(os.path.join(excelPath, excelFile))
        for sheetname in wb.sheetnames:
            print('>>> Poccessing sheet[' + sheetname + ']')
            ws = wb[sheetname]
            csvfilename = excelFile + '-' + sheetname + '.csv'
            csvFilePath = os.path.join(csvPath, csvfilename)
            csvFileObj = open(csvFilePath, 'w', newline='')
            writer = csv.writer(csvFileObj)
            for rowNum in range(1, ws.max_row+1):
                rowData = []
                for colNum in range(1, ws.max_column+1):
                    rowData.append(ws.cell(rowNum, colNum).value)
                writer.writerow(rowData)
            csvFileObj.close()
            print('>>> Completed to generate csv file: ', csvFilePath)
    print('Done!')

def main():
    csvToExcel()

if __name__ == '__main__':
    main()